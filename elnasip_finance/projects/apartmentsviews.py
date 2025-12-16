# projects/views.py
from django.shortcuts import render
from django.db.models import Q, Sum, Count
from .models import Apartment, Block, Project
from projects.apartments.forms import ApartmentFilterForm
# projects/views.py
import csv
from django.http import HttpResponse
from django.shortcuts import render
from django.db.models import Q, Sum, Count
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

def export_apartments_xlsx(qs):
    wb = Workbook()
    ws = wb.active
    ws.title = "Квартиры"

    headers = [
        'ЖК',
        'Блок',
        'Этаж',
        'Номер',
        'Комнат',
        'Площадь, m²',
        'План. цена m²',
        'Факт. цена m²',
        'Поступило со сделки, сом',
        'Цена сделки, сом',
        'Остаток по сделке, сом',
        'Статус',
        'Клиент / арендатор',
        'Телефон арендатора',
        'Номер сделки',
        'Договор аренды',
    ]

    # Заголовки
    ws.append(headers)
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Данные
    for a in qs:
        if a.is_sold:
            status = 'Продана'
        elif a.is_reserved:
            status = 'Бронь'
        elif a.is_rented:
            status = 'Аренда'
        else:
            status = 'Свободна'

        client_or_tenant = a.client_name or a.tenant_name or ''

        ws.append([
            a.block.project.name if a.block and a.block.project else '',
            a.block.name if a.block else '',
            a.floor,
            a.apartment_number,
            a.rooms,
            float(a.area or 0),
            float(a.planned_price_per_m2 or 0),
            float(a.fact_price_per_m2 or 0),
            float(a.deal_amount or 0),
            float(a.deal_Fakt_deal_amount or 0),
            float(a.remaining_deal_amount or 0),
            status,
            client_or_tenant,
            a.tenant_phone or '',
            a.deal_number or '',
            a.tenant_contract or '',
        ])

    # Форматы колонок (как числа)
    num_cols_2dec = [6, 7, 8]            # m2, цены за м2
    money_cols_0dec = [9, 10, 11]        # деньги
    int_cols = [3, 5]                    # этаж, комнаты

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in int_cols:
            row[c-1].number_format = '0'
        for c in num_cols_2dec:
            row[c-1].number_format = '#,##0.00'
        for c in money_cols_0dec:
            row[c-1].number_format = '#,##0'

    # Авто-ширина колонок
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[letter].width = min(max_len + 2, 45)

    # Ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="apartments.xlsx"'
    wb.save(response)
    return response

def apartment_list(request):
    qs = Apartment.objects.select_related(
        'block',
        'block__project',
    ).all()

    form = ApartmentFilterForm(request.GET or None)

    # Динамический queryset для блоков
    project_id = request.GET.get('project')
    if project_id:
        form.fields['block'].queryset = Block.objects.filter(project_id=project_id)
    else:
        form.fields['block'].queryset = Block.objects.all()

    if form.is_valid():
        cd = form.cleaned_data

        # ЖК
        if cd.get('project'):
            qs = qs.filter(block__project=cd['project'])

        # Блок
        if cd.get('block'):
            qs = qs.filter(block=cd['block'])

        # Статус
        status = cd.get('status')
        if status == 'free':
            qs = qs.filter(is_sold=False, is_reserved=False, is_rented=False)
        elif status == 'sold':
            qs = qs.filter(is_sold=True)
        elif status == 'reserved':
            qs = qs.filter(is_reserved=True, is_sold=False)
        elif status == 'rented':
            qs = qs.filter(is_rented=True)

        # Комнаты
        if cd.get('rooms_min') is not None:
            qs = qs.filter(rooms__gte=cd['rooms_min'])
        if cd.get('rooms_max') is not None:
            qs = qs.filter(rooms__lte=cd['rooms_max'])

        # Площадь
        if cd.get('area_min') is not None:
            qs = qs.filter(area__gte=cd['area_min'])
        if cd.get('area_max') is not None:
            qs = qs.filter(area__lte=cd['area_max'])

        # Планируемая цена m²
        if cd.get('price_min') is not None:
            qs = qs.filter(planned_price_per_m2__gte=cd['price_min'])
        if cd.get('price_max') is not None:
            qs = qs.filter(planned_price_per_m2__lte=cd['price_max'])

        # Этаж
        if cd.get('floor_min') is not None:
            qs = qs.filter(floor__gte=cd['floor_min'])
        if cd.get('floor_max') is not None:
            qs = qs.filter(floor__lte=cd['floor_max'])

        # Поиск по клиенту / арендатору / договору
        search = cd.get('client_search')
        if search:
            qs = qs.filter(
                Q(client_name__icontains=search) |
                Q(tenant_name__icontains=search) |
                Q(tenant_phone__icontains=search) |
                Q(deal_number__icontains=search) |
                Q(tenant_contract__icontains=search)
            )

        # Сортировка
        order = cd.get('order')
        if order:
            if order == '':
                qs = qs.order_by(
                    'block__project__name',
                    'block__name',
                    'floor',
                    'apartment_number',
                )
            else:
                qs = qs.order_by(order)
        else:
            qs = qs.order_by(
                'block__project__name',
                'block__name',
                'floor',
                'apartment_number',
            )
    else:
        qs = qs.order_by(
            'block__project__name',
            'block__name',
            'floor',
            'apartment_number',
        )

    # 👉 Если нажали "Экспорт в Excel" — сразу отдаем файл
    if request.GET.get('export') == 'xlsx':
        return export_apartments_xlsx(qs)

    # Статистика для шапки
    stats = qs.aggregate(
        total=Count('id'),
        sold=Count('id', filter=Q(is_sold=True)),
        free=Count('id', filter=Q(is_sold=False, is_reserved=False, is_rented=False)),
        total_area=Sum('area'),
        sold_area=Sum('sold_area'),
    )

    context = {
        'form': form,
        'apartments': qs,
        'stats': stats,
    }
    return render(request, 'projects/apartments_list.html', context)


