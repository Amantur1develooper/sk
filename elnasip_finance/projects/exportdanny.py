import re
from decimal import Decimal, InvalidOperation
from datetime import datetime, date, time

from django.contrib import messages
from django.db import transaction
from django.db.models import Sum
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.utils import timezone

from openpyxl import load_workbook
import xlrd

from .models import Block, Apartment, DealPayment, ApartmentComment
from .forms import BlockApartmentsImportForm
from finances.models import CommonCash, CashFlow, Sale


class _XlsSheet:
    """Обёртка над xlrd.Sheet, имитирующая нужный интерфейс openpyxl-листа."""

    def __init__(self, sheet):
        self._sheet = sheet

    @property
    def max_column(self):
        return self._sheet.ncols

    @property
    def max_row(self):
        return self._sheet.nrows

    def cell(self, row, column):
        return _XlsCell(self._sheet, row - 1, column - 1)


class _XlsCell:
    """Обёртка над значением ячейки xlrd, имитирующая openpyxl Cell."""

    def __init__(self, sheet, row0, col0):
        self._sheet = sheet
        self._row0 = row0
        self._col0 = col0

    @property
    def value(self):
        if self._row0 >= self._sheet.nrows or self._col0 >= self._sheet.ncols:
            return None
        cell = self._sheet.cell(self._row0, self._col0)
        ctype = cell.ctype
        val = cell.value
        # xlrd.XL_CELL_DATE -> конвертируем в datetime
        if ctype == xlrd.XL_CELL_DATE:
            try:
                tup = xlrd.xldate_as_tuple(val, self._sheet.book.datemode)
                return datetime(*tup) if tup[0] else None
            except Exception:
                return val
        # пустая ячейка
        if ctype == xlrd.XL_CELL_EMPTY:
            return None
        # число — оставляем как есть (float/int)
        return val


def _load_sheet(file_obj, filename):
    """Возвращает объект листа (openpyxl или _XlsSheet) в зависимости от расширения."""
    name_lower = filename.lower()
    if name_lower.endswith(".xlsx"):
        wb = load_workbook(file_obj, data_only=True)
        return wb.active
    elif name_lower.endswith(".xls"):
        data = file_obj.read()
        wb = xlrd.open_workbook(file_contents=data)
        sheet = wb.sheet_by_index(0)
        sheet.book = wb  # нужно для datemode в _XlsCell
        return _XlsSheet(sheet)
    else:
        return None


def _to_decimal(val):
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        val = val.replace(" ", "").replace(",", ".")
    try:
        return Decimal(str(val))
    except (InvalidOperation, ValueError):
        return None


def _is_empty_cell(val):
    return val is None or (isinstance(val, str) and val.strip() == "")


def _normalize_name(raw):
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    if not re.search(r"[A-Za-zА-Яа-яЁё]", s):
        return ""
    s = re.sub(r"\s+", " ", s).strip(" ,.;:-")
    return s


def _to_int(val, default=0):
    if val is None:
        return default
    try:
        if isinstance(val, str):
            val = val.strip()
            if not val:
                return default
        return int(float(val))
    except (ValueError, TypeError):
        return default


def _clean_apartment_number(raw):
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    return s.split("(")[0].strip()


def _parse_contract_date(val):
    if val is None:
        return timezone.now()

    if isinstance(val, datetime):
        dt = val
    elif isinstance(val, date):
        dt = datetime.combine(val, time(0, 0))
    elif isinstance(val, str):
        s = val.strip()
        dt = None
        for fmt in ("%d-%m-%Y", "%d.%m.%Y", "%Y-%m-%d"):
            try:
                dt = datetime.strptime(s, fmt)
                break
            except ValueError:
                pass
        if dt is None:
            return timezone.now()
    else:
        return timezone.now()

    if timezone.is_naive(dt):
        return timezone.make_aware(dt, timezone.get_current_timezone())
    return dt


def _deal_type_flags(raw_type):
    """Определяет is_barter, is_reserved по колонке 'Тип сделки'."""
    s = str(raw_type or "").strip().lower()
    is_barter = "бартер" in s
    is_reserved = ("брон" in s) and not is_barter
    return is_barter, is_reserved


@login_required
def import_block_apartments_excel(request, block_id):
    block = get_object_or_404(Block, id=block_id)

    if request.method == "GET":
        form = BlockApartmentsImportForm()
        return render(request, "projects/block_apartments_import.html", {"block": block, "form": form})

    form = BlockApartmentsImportForm(request.POST, request.FILES)
    if not form.is_valid():
        return render(request, "projects/block_apartments_import.html", {"block": block, "form": form})

    uploaded = form.cleaned_data["file"]
    update_existing = form.cleaned_data.get("update_existing", True)
    create_missing = form.cleaned_data.get("create_missing", True)

    ws = _load_sheet(uploaded, uploaded.name)
    if ws is None:
        messages.error(request, "Поддерживаются только файлы .xlsx и .xls.")
        return render(request, "projects/block_apartments_import.html", {"block": block, "form": form})

    # --- читаем заголовки ---
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[str(v).strip()] = c

    required = ["Предварительный номер", "Количество комнат", "Общая площадь ориентировочно"]
    miss = [h for h in required if h not in headers]
    if miss:
        messages.error(request, f"В файле не хватает колонок: {', '.join(miss)}")
        return render(request, "projects/block_apartments_import.html", {"block": block, "form": form})

    def col(name):
        """Возвращает значение ячейки по имени колонки или None."""
        if name not in headers:
            return None
        return ws.cell(row=row, column=headers[name]).value

    created_cnt = 0
    updated_cnt = 0
    skipped_cnt = 0
    returns_cnt = 0      # квартира была продана, в Excel — пустая (возврат)
    resales_cnt = 0      # квартира продана другому клиенту (перепродажа)
    barter_cnt = 0       # бартерные сделки
    errors = []
    event_log = []       # детальный лог изменений

    for row in range(2, ws.max_row + 1):
        try:
            raw_num = col("Предварительный номер")
            apt_number = _clean_apartment_number(raw_num)
            if not apt_number:
                skipped_cnt += 1
                continue

            # --- читаем все поля из Excel ---
            fio = _normalize_name(col("Ф.И.О."))
            note_raw = col("Примечание") or ""
            phone_raw = col("Телефон")
            phone = str(phone_raw).strip() if phone_raw else ""

            deal_type_raw = col("Тип сделки") or ""
            is_barter, is_reserved_from_type = _deal_type_flags(deal_type_raw)

            paid_cell = col("Оплачено")
            paid = _to_decimal(paid_cell) or Decimal("0")

            sum_contract = _to_decimal(col("Сумма договора")) or Decimal("0")
            remaining_excel = _to_decimal(col("Остаток на оплату"))
            if remaining_excel is None:
                remaining_excel = max(sum_contract - paid, Decimal("0"))

            price_m2 = _to_decimal(col("Цена за 1 м.кв."))
            floor = _to_int(col("Этаж"), default=1)
            rooms = _to_int(col("Количество комнат"), default=0)
            area = _to_decimal(col("Общая площадь ориентировочно")) or Decimal("0")
            contract_dt = _parse_contract_date(col("Дата договора"))

            # --- определяем статус из Excel ---
            excel_has_client = bool(fio)
            # Продана если: есть ФИО, или оплачено > 0, или бартер с суммой договора
            excel_is_sold = excel_has_client or paid > 0 or (is_barter and sum_contract > 0)
            # Только бронь (не продана)
            is_reserved = is_reserved_from_type and not excel_is_sold

            # --- ищем в БД ---
            apt_existing = Apartment.objects.filter(block=block, apartment_number=apt_number).first()
            exists = apt_existing is not None

            if exists and not update_existing:
                skipped_cnt += 1
                continue
            if not exists and not create_missing:
                skipped_cnt += 1
                continue

            # --- детектим особые случаи ---
            event_comment = None
            event_tag = None

            if exists and apt_existing.is_sold:
                db_client = (apt_existing.client_name or "").strip()

                if not excel_is_sold and not is_reserved:
                    # ВОЗВРАТ: в Excel пусто, в БД — продана
                    event_tag = "ВОЗВРАТ"
                    event_comment = f"[Импорт] Возврат: ранее продана клиенту «{db_client}». Excel показывает квартиру свободной."
                    returns_cnt += 1

                elif excel_has_client and db_client:
                    norm_db = re.sub(r"\s+", " ", db_client).lower()
                    norm_xl = re.sub(r"\s+", " ", fio).lower()
                    if norm_db != norm_xl:
                        # ПЕРЕПРОДАЖА: клиент изменился
                        event_tag = "ПЕРЕПРОДАЖА"
                        event_comment = f"[Импорт] Перепродажа: «{db_client}» → «{fio}»."
                        resales_cnt += 1

            if is_barter:
                barter_cnt += 1
                if event_tag is None:
                    event_tag = "БАРТЕР"

            # --- формируем defaults для update_or_create ---
            defaults = {
                "floor": floor,
                "rooms": rooms,
                "area": area,
                "is_sold": excel_is_sold,
                "is_reserved": is_reserved,
                "is_barter": is_barter,
                "sold_area": Decimal("0"),  # не пересчитываем — суммы идут из Excel
            }

            if excel_has_client:
                defaults["client_name"] = fio
            elif event_tag == "ВОЗВРАТ":
                defaults["client_name"] = None
                defaults["is_sold"] = False
                defaults["is_reserved"] = False

            if price_m2 is not None:
                defaults["fact_price_per_m2"] = price_m2
                defaults["planned_price_per_m2"] = price_m2

            with transaction.atomic():
                apt, created_flag = Apartment.objects.update_or_create(
                    block=block,
                    apartment_number=apt_number,
                    defaults=defaults,
                )

                # --- финансовые поля напрямую (обходим auto-recalc в save()) ---
                Apartment.objects.filter(id=apt.id).update(
                    deal_Fakt_deal_amount=sum_contract,
                    deal_amount=paid,
                    remaining_deal_amount=remaining_excel,
                    planned_deal_amount=Decimal("0") if excel_is_sold else (area * price_m2 if price_m2 and area else Decimal("0")),
                )

                if created_flag:
                    created_cnt += 1
                else:
                    updated_cnt += 1

                # --- комментарий к квартире ---
                comment_parts = []
                if event_comment:
                    comment_parts.append(event_comment)
                note_str = str(note_raw).strip() if note_raw else ""
                if note_str:
                    comment_parts.append(f"Примечание: {note_str[:300]}")
                if phone:
                    comment_parts.append(f"Телефон: {phone}")
                if is_barter and "бартер" not in note_str.lower():
                    comment_parts.append(f"Тип сделки: Бартер ({str(deal_type_raw).strip()})")

                if comment_parts:
                    full_text = "[Импорт Excel] " + " | ".join(comment_parts)
                    # Не дублируем комментарии
                    if not ApartmentComment.objects.filter(apartment=apt, text=full_text).exists():
                        ApartmentComment.objects.create(
                            apartment=apt,
                            text=full_text,
                            author=request.user.get_username(),
                        )

                if event_tag:
                    event_log.append(f"{apt_number}: {event_tag}" + (f" ({fio or 'свободна'})" if event_tag != "БАРТЕР" else ""))

        except Exception as e:
            errors.append(f"Строка {row} ({apt_number if apt_number else '?'}): {e}")

    messages.success(
        request,
        f"Импорт завершён. Создано: {created_cnt}, обновлено: {updated_cnt}, "
        f"пропущено: {skipped_cnt}. Возвратов: {returns_cnt}, перепродаж: {resales_cnt}, бартеров: {barter_cnt}."
    )

    return render(request, "projects/block_apartments_import.html", {
        "block": block,
        "form": BlockApartmentsImportForm(),
        "created": created_cnt,
        "updated": updated_cnt,
        "skipped": skipped_cnt,
        "returns": returns_cnt,
        "resales": resales_cnt,
        "barters": barter_cnt,
        "event_log": event_log,
        "errors": errors,
    })



from datetime import datetime
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import Block, Apartment


def _dec(val, default=Decimal("0")):
    try:
        return Decimal(val or 0)
    except Exception:
        return default


def _safe_str(val):
    return "" if val is None else str(val).strip()


@login_required
def export_block_apartments_excel(request, block_id):
    """
    Экспорт квартир конкретного блока в .xlsx
    Формат заголовков совместим с import_block_apartments_excel().
    Поддерживает простые фильтры через GET:
      ?only_free=1  -> только свободные
      ?only_rented=1 -> только аренда
      ?only_sold=1  -> только проданные
      ?only_reserved=1 -> только бронь (не проданные)
    """
    block = get_object_or_404(Block, id=block_id)

    qs = Apartment.objects.filter(block=block).prefetch_related("payments", "comments").order_by("floor", "apartment_number")

    # --- фильтры (если нужны) ---
    if request.GET.get("only_free") == "1":
        qs = qs.filter(is_sold=False, is_reserved=False, is_rented=False)
    if request.GET.get("only_rented") == "1":
        qs = qs.filter(is_rented=True)
    if request.GET.get("only_sold") == "1":
        qs = qs.filter(is_sold=True)
    if request.GET.get("only_reserved") == "1":
        qs = qs.filter(is_reserved=True, is_sold=False)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Блок {block.name}"

    headers = [
        "Этаж",
        "Предварительный номер",
        "Количество комнат",
        "Общая площадь ориентировочно",
        "Дата договора",
        "Ф.И.О.",
        "Цена за 1 м.кв.",
        "Сумма договора",
        "Оплачено",
        "Остаток на оплату",
        "Телефон",
        "Примечание",
    ]
    ws.append(headers)

    # ширины колонок (чтобы было удобно)
    widths = [7, 22, 18, 26, 16, 28, 16, 16, 12, 18, 16, 35]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for apt in qs:
        paid_total = apt.payments.aggregate(total=Sum("amount"))["total"] or Decimal("0")  # важно: как в импорте сравниваешь с платежами
        paid_total = _dec(paid_total)

        contract_sum = _dec(apt.deal_Fakt_deal_amount, default=Decimal("0"))
        if contract_sum <= 0:
            # если суммы договора нет — можно подставить планируемую
            contract_sum = _dec(apt.planned_deal_amount, default=Decimal("0"))

        remaining = _dec(apt.remaining_deal_amount, default=(contract_sum - paid_total))
        if remaining < 0:
            remaining = Decimal("0")

        # дата договора: у тебя явного поля нет — экспортируем created_at (стабильно для обратной загрузки)
        contract_date = apt.created_at.date() if getattr(apt, "created_at", None) else timezone.localdate()

        fio = _safe_str(apt.client_name)

        # телефон: отдельного поля телефона клиента нет — оставим пусто (или можешь заменить на своё поле)
        phone = ""

        # примечание: если бронь — пишем "бронь", чтобы импорт корректно понял
        note_parts = []
        if apt.is_reserved and not apt.is_sold:
            note_parts.append("бронь")

        # (опционально) подтянуть последний комментарий
        last_comment = apt.comments.order_by("-created_at").first()
        if last_comment and last_comment.text:
            txt = last_comment.text.replace("[Импорт Excel]", "").strip()
            if txt:
                note_parts.append(txt[:200])

        note = " | ".join(note_parts)

        ws.append([
            int(apt.floor or 1),
            _safe_str(apt.apartment_number),
            int(apt.rooms or 0),
            float(_dec(apt.area)),
            contract_date.strftime("%d.%m.%Y"),
            fio,
            float(_dec(apt.planned_price_per_m2)),
            float(contract_sum),
            float(paid_total),
            float(remaining),
            phone,
            note,
        ])

    filename = f"block_{block.id}_{block.name}_{timezone.localdate().isoformat()}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
