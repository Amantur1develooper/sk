from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum, Avg, Count
from django.utils import timezone
from datetime import datetime, timedelta
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from finances.models import CommonCash, CashFlow, Allocation, Expense, Sale
from projects.models import Project, Block, EstimateItem
from employees.models import Employee, SalaryPayment

@login_required
def reports_dashboard(request):
    """Главная страница отчетов"""
    return render(request, 'reports/reports_dashboard.html')

@login_required
def cash_flow_report(request):
    """Отчет по движению денег в Общаге"""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    cash_flows = CashFlow.objects.all().order_by('-date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        cash_flows = cash_flows.filter(date__gte=start_date)
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        cash_flows = cash_flows.filter(date__lte=end_date)
    
    # Суммируем приходы и расходы
    total_income = cash_flows.filter(flow_type='income').aggregate(Sum('amount'))['amount__sum'] or 0
    total_expense = cash_flows.filter(flow_type='expense').aggregate(Sum('amount'))['amount__sum'] or 0
    balance = total_income - total_expense
    
    context = {
        'cash_flows': cash_flows,
        'total_income': total_income,
        'total_expense': total_expense,
        'balance': balance,
        'start_date': start_date.strftime('%Y-%m-%d') if start_date else '',
        'end_date': end_date.strftime('%Y-%m-%d') if end_date else '',
    }
    
    if 'export' in request.GET:
        return export_cash_flow_report(cash_flows, total_income, total_expense, balance)
    
    return render(request, 'reports/cash_flow_report.html', context)

@login_required
def project_report(request, project_id=None):
    """Отчет по проекту/объекту"""
    projects = Project.objects.all()
    selected_project = None
    blocks = None
    project_data = []
    
    if project_id:
        selected_project = Project.objects.get(id=project_id)
        blocks = selected_project.blocks.all()
        
        for block in blocks:
            # Собираем данные по каждому блоку
            planned = sum(item.planned_amount for item in block.estimate_items.all())
            allocated = sum(item.get_allocated_sum() for item in block.estimate_items.all())
            spent = sum(item.spent_amount for item in block.estimate_items.all())
            margin = planned - spent
            
            project_data.append({
                'block': block,
                'planned': planned,
                'allocated': allocated,
                'spent': spent,
                'margin': margin,
                'sold_area': block.sold_area,
                'remaining_area': block.remaining_area,
            })
    
    context = {
        'projects': projects,
        'selected_project': selected_project,
        'project_data': project_data,
    }
    
    if 'export' in request.GET and selected_project:
        return export_project_report(selected_project, project_data)
    
    return render(request, 'reports/project_report.html', context)

@login_required
def block_report(request, block_id=None):
    """Отчет по блоку с детализацией по статьям сметы"""
    blocks = Block.objects.all()
    selected_block = None
    estimate_data = []
    
    if block_id:
        selected_block = Block.objects.get(id=block_id)
        estimate_items = selected_block.estimate_items.all()
        
        for item in estimate_items:
            allocated = item.get_allocated_sum()
            spent = item.spent_amount
            margin = item.planned_amount - spent
            
            estimate_data.append({
                'item': item,
                'planned': item.planned_amount,
                'allocated': allocated,
                'spent': spent,
                'margin': margin,
                'completion': (spent / item.planned_amount * 100) if item.planned_amount > 0 else 0,
            })
    
    context = {
        'blocks': blocks,
        'selected_block': selected_block,
        'estimate_data': estimate_data,
    }
    
    if 'export' in request.GET and selected_block:
        return export_block_report(selected_block, estimate_data)
    
    return render(request, 'reports/block_report.html', context)

@login_required
def allocations_report(request):
    """Отчет по выделениям средств"""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    project_id = request.GET.get('project_id')
    
    allocations = Allocation.objects.all().select_related(
        'estimate_item', 'estimate_item__block', 'estimate_item__block__project'
    ).order_by('-date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        allocations = allocations.filter(date__gte=start_date)
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        allocations = allocations.filter(date__lte=end_date)
    
    if project_id:
        allocations = allocations.filter(estimate_item__block__project_id=project_id)
    
    total_amount = allocations.aggregate(Sum('amount'))['amount__sum'] or 0
    
    context = {
        'allocations': allocations,
        'total_amount': total_amount,
        'projects': Project.objects.all(),
        'start_date': start_date.strftime('%Y-%m-%d') if start_date else '',
        'end_date': end_date.strftime('%Y-%m-%d') if end_date else '',
        'selected_project': project_id,
    }
    
    if 'export' in request.GET:
        return export_allocations_report(allocations, total_amount)
    
    return render(request, 'reports/allocations_report.html', context)

@login_required
def salary_report(request):
    """Отчет по зарплатам"""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    payments = SalaryPayment.objects.all().select_related('employee', 'employee__user').order_by('-date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        payments = payments.filter(date__gte=start_date)
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        payments = payments.filter(date__lte=end_date)
    
    total_amount = payments.aggregate(Sum('amount'))['amount__sum'] or 0
    salary_count = payments.filter(payment_type='salary').count()
    advance_count = payments.filter(payment_type='advance').count()
    
    context = {
        'payments': payments,
        'total_amount': total_amount,
        'salary_count': salary_count,
        'advance_count': advance_count,
        'start_date': start_date.strftime('%Y-%m-%d') if start_date else '',
        'end_date': end_date.strftime('%Y-%m-%d') if end_date else '',
    }
    
    if 'export' in request.GET:
        return export_salary_report(payments, total_amount, salary_count, advance_count)
    
    return render(request, 'reports/salary_report.html', context)

@login_required
def sales_report(request):
    """Отчет по продажам"""
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    project_id = request.GET.get('project_id')
    
    sales = Sale.objects.all().select_related('block', 'block__project').order_by('-date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        sales = sales.filter(date__gte=start_date)
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        sales = sales.filter(date__lte=end_date)
    
    if project_id:
        sales = sales.filter(block__project_id=project_id)
    
    total_amount = sales.aggregate(Sum('amount'))['amount__sum'] or 0
    total_area = sales.aggregate(Sum('area'))['area__sum'] or 0
    avg_price = total_amount / total_area if total_area > 0 else 0
    
    context = {
        'sales': sales,
        'total_amount': total_amount,
        'total_area': total_area,
        'avg_price': avg_price,
        'projects': Project.objects.all(),
        'start_date': start_date.strftime('%Y-%m-%d') if start_date else '',
        'end_date': end_date.strftime('%Y-%m-%d') if end_date else '',
        'selected_project': project_id,
    }
    
    if 'export' in request.GET:
        return export_sales_report(sales, total_amount, total_area, avg_price)
    
    return render(request, 'reports/sales_report.html', context)

# Функции экспорта в Excel
def export_cash_flow_report(cash_flows, total_income, total_expense, balance):
    """Экспорт отчета по движению денег в Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Движение денежных средств"
    
    # Заголовки
    headers = ['Дата', 'Тип операции', 'Сумма', 'Описание', 'Создано']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header).font = Font(bold=True)
    
    # Данные
    for row, flow in enumerate(cash_flows, 2):
        ws.cell(row=row, column=1, value=flow.date.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row, column=2, value=flow.get_flow_type_display())
        ws.cell(row=row, column=3, value=float(flow.amount))
        ws.cell(row=row, column=4, value=flow.description)
        ws.cell(row=row, column=5, value=flow.created_by.username)
    
    # Итоги
    ws.cell(row=row+2, column=2, value="Общий приход:").font = Font(bold=True)
    ws.cell(row=row+2, column=3, value=float(total_income)).font = Font(bold=True)
    
    ws.cell(row=row+3, column=2, value="Общий расход:").font = Font(bold=True)
    ws.cell(row=row+3, column=3, value=float(total_expense)).font = Font(bold=True)
    
    ws.cell(row=row+4, column=2, value="Баланс:").font = Font(bold=True)
    ws.cell(row=row+4, column=3, value=float(balance)).font = Font(bold=True)
    
    # Форматирование
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем файл
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="cash_flow_report.xlsx"'
    wb.save(response)
    
    return response

# Аналогичные функции экспорта для других отчетов
def export_project_report(project, project_data):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Отчет по проекту {project.name}"
    
    headers = ['Блок', 'Плановая сумма', 'Выделено средств', 'Факт расходов', 'Маржа', 'Проданная площадь', 'Остаток площади']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header).font = Font(bold=True)
    
    for row, data in enumerate(project_data, 2):
        ws.cell(row=row, column=1, value=data['block'].name)
        ws.cell(row=row, column=2, value=float(data['planned']))
        ws.cell(row=row, column=3, value=float(data['allocated']))
        ws.cell(row=row, column=4, value=float(data['spent']))
        ws.cell(row=row, column=5, value=float(data['margin']))
        ws.cell(row=row, column=6, value=float(data['sold_area']))
        ws.cell(row=row, column=7, value=float(data['remaining_area']))
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="project_report_{project.name}.xlsx"'
    wb.save(response)
    
    return response

# Другие функции экспорта (allocations, salary, sales) реализуются аналогично

def export_allocations_report(allocations, total_amount):
    # Реализация аналогична export_cash_flow_report
    pass

def export_salary_report(payments, total_amount, salary_count, advance_count):
    # Реализация аналогична export_cash_flow_report
    pass

def export_sales_report(sales, total_amount, total_area, avg_price):
    # Реализация аналогична export_cash_flow_report
    pass