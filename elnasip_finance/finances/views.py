from django.shortcuts import render, get_object_or_404

from .models import CommonCash, CashFlow, Allocation, Expense, Sale
from projects.models import Project, Block, EstimateItem
from employees.models import Employee, SalaryPayment
from django.db.models import Sum
from django.db.models import Sum, Q
from django.db.models import Q
from django.utils.dateparse import parse_date
from django.http import JsonResponse
from django.db.models import Q
from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from .models import WarehouseCar
from .forms import CarPurchaseForm, CarSaleForm
from .models import Block, EstimateItem
from .forms import AllocationForm

@login_required
def common_cash_detail(request):
    common_cash = CommonCash.objects.first()
    
    # Фильтрация для движения денег
    cash_flows = CashFlow.objects.all().order_by('-date')
    
    # Фильтры
    flow_type = request.GET.get('flow_type')
    block_id = request.GET.get('block')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if flow_type:
        cash_flows = cash_flows.filter(flow_type=flow_type)
    
    if block_id:
        cash_flows = cash_flows.filter(block_id=block_id)
    
    if start_date:
        cash_flows = cash_flows.filter(date__gte=start_date)
    
    if end_date:
        cash_flows = cash_flows.filter(date__lte=end_date)
    
    # Получаем все выделения средств
    allocations = Allocation.objects.all().select_related(
        'estimate_item', 'estimate_item__block'
    ).order_by('-date')
    
    # Статистика по выделениям
    total_allocated = allocations.aggregate(Sum('amount'))['amount__sum'] or 0
    average_allocation = total_allocated / allocations.count() if allocations.count() > 0 else 0
    
    # Статистика по движению денег
    total_income = cash_flows.filter(flow_type='income').aggregate(Sum('amount'))['amount__sum'] or 0
    total_expense = cash_flows.filter(flow_type='expense').aggregate(Sum('amount'))['amount__sum'] or 0
    net_balance = total_income - total_expense
    
    context = {
        'common_cash': common_cash,
        'cash_flows': cash_flows,
        'allocations': allocations,
        'total_income': total_income,
        'total_expense': total_expense,
        'net_balance': net_balance,
        'total_allocated': total_allocated,
        'average_allocation': average_allocation,
        'blocks': Block.objects.all(),
    }
    
    
    return render(request, 'finances/common_cash_detail.html', context)

@login_required
def common_cash_detail2(request):
    common_cash = CommonCash.objects.first()
    
    # Фильтрация для движения денег
    cash_flows = CashFlow.objects.all().order_by('-date')
    
    # Фильтры
    flow_type = request.GET.get('flow_type')
    block_id = request.GET.get('block')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if flow_type:
        cash_flows = cash_flows.filter(flow_type=flow_type)
    
    if block_id:
        cash_flows = cash_flows.filter(block_id=block_id)
    
    if start_date:
        cash_flows = cash_flows.filter(date__gte=start_date)
    
    if end_date:
        cash_flows = cash_flows.filter(date__lte=end_date)
    
    # Получаем все выделения средств
    allocations = Allocation.objects.all().select_related(
        'estimate_item', 'estimate_item__block'
    ).order_by('-date')
    
    # Статистика по выделениям
    total_allocated = allocations.aggregate(Sum('amount'))['amount__sum'] or 0
    average_allocation = total_allocated / allocations.count() if allocations.count() > 0 else 0
    
    # Статистика по движению денег
    total_income = cash_flows.filter(flow_type='income').aggregate(Sum('amount'))['amount__sum'] or 0
    total_expense = cash_flows.filter(flow_type='expense').aggregate(Sum('amount'))['amount__sum'] or 0
    net_balance = total_income - total_expense
    
    context = {
        'common_cash': common_cash,
        'cash_flows': cash_flows,
        'allocations': allocations,
        'total_income': total_income,
        'total_expense': total_expense,
        'net_balance': net_balance,
        'total_allocated': total_allocated,
        'average_allocation': average_allocation,
        'blocks': Block.objects.all(),
    }
    
    
    return render(request, 'finances/common_cash_detail2.html', context)







@login_required
def dashboard(request):
    import json
    from django.db.models import Sum, Q
    from projects.models import Project, Block, Apartment

    common_cash = CommonCash.objects.first()
    projects = Project.objects.prefetch_related('blocks').all()
    total_projects = projects.count()
    total_blocks = Block.objects.count()

    total_income  = CashFlow.objects.filter(flow_type='income').aggregate(s=Sum('amount'))['s'] or 0
    total_expense = CashFlow.objects.filter(flow_type='expense').aggregate(s=Sum('amount'))['s'] or 0
    margin = total_income - total_expense

    refund_total  = CashFlow.objects.filter(description__startswith="ВОЗВРАТ:").aggregate(s=Sum('amount'))['s'] or 0
    refund_total2 = CashFlow.objects.filter(description__startswith="УДАЛЕНИЕ:").aggregate(s=Sum('amount'))['s'] or 0
    total_income2  = total_income  - refund_total
    total_expense2 = total_expense - refund_total2
    margin_mini = total_income2 - total_expense2

    recent_operations = CashFlow.objects.select_related('block__project').order_by('-date')[:15]

    # ── Приход по ЖК ──────────────────────────────────────────────────────
    income_by_project = []
    for proj in projects:
        amt = CashFlow.objects.filter(
            flow_type='income', block__project=proj
        ).aggregate(s=Sum('amount'))['s'] or 0
        # также прямые платежи без block (например, импорт) через Apartment
        if amt:
            income_by_project.append({'name': proj.name, 'amount': float(amt)})

    # fallback — если block не привязан, считаем через Apartment.deal_amount
    project_income_labels = [p['name'] for p in income_by_project]
    project_income_data   = [p['amount'] for p in income_by_project]

    # ── Приход по блокам (топ-15) ─────────────────────────────────────────
    blocks_income_qs = (
        CashFlow.objects
        .filter(flow_type='income', block__isnull=False)
        .values('block__name', 'block__project__name')
        .annotate(s=Sum('amount'))
        .order_by('-s')[:15]
    )
    block_income_labels = [f"{r['block__project__name']} / {r['block__name']}" for r in blocks_income_qs]
    block_income_data   = [float(r['s']) for r in blocks_income_qs]

    # ── Куда уходят деньги (расходы по описанию — первые слова) ──────────
    expense_rows = (
        CashFlow.objects
        .filter(flow_type='expense')
        .values('description')
        .annotate(s=Sum('amount'))
    )
    expense_map = {}
    for row in expense_rows:
        desc = row['description'] or ''
        # первые 2 слова как категория
        key = ' '.join(desc.split()[:2]) or 'Прочее'
        # группируем по смысловым блокам
        if 'Выделение' in key or 'выдел' in desc.lower():
            cat = 'Смета/выделения'
        elif any(w in desc for w in ('Зарплата', 'Аванс', 'Бонус', 'Премия')):
            cat = 'Зарплата и бонусы'
        elif 'займ' in desc.lower() or 'Займ' in desc:
            cat = 'Займы'
        else:
            cat = 'Прочие расходы'
        expense_map[cat] = expense_map.get(cat, 0) + float(row['s'])
    expense_labels = list(expense_map.keys())
    expense_data   = list(expense_map.values())

    # ── Динамика приход/расход по месяцам (последние 12) ─────────────────
    from django.db.models.functions import TruncMonth
    monthly = (
        CashFlow.objects
        .annotate(month=TruncMonth('date'))
        .values('month', 'flow_type')
        .annotate(s=Sum('amount'))
        .order_by('month')
    )
    months_map = {}
    for row in monthly:
        m = row['month'].strftime('%b %Y') if row['month'] else '?'
        if m not in months_map:
            months_map[m] = {'income': 0, 'expense': 0}
        months_map[m][row['flow_type']] = float(row['s'])
    # последние 12
    all_months = list(months_map.items())[-12:]
    monthly_labels  = [m for m, _ in all_months]
    monthly_income  = [v['income']  for _, v in all_months]
    monthly_expense = [v['expense'] for _, v in all_months]

    # ── Статистика квартир ─────────────────────────────────────────────────
    apt_total    = Apartment.objects.count()
    apt_sold     = Apartment.objects.filter(is_sold=True).count()
    apt_reserved = Apartment.objects.filter(is_reserved=True, is_sold=False).count()
    apt_barter   = Apartment.objects.filter(is_barter=True).count()
    apt_free     = Apartment.objects.filter(is_sold=False, is_reserved=False, is_barter=False).count()

    context = {
        'common_cash': common_cash,
        'projects': projects,
        'total_projects': total_projects,
        'total_blocks': total_blocks,
        'total_income': total_income,
        'total_income2': total_income2,
        'total_expense': total_expense,
        'total_expense2': total_expense2,
        'margin': margin,
        'margin_mini': margin_mini,
        'recent_operations': recent_operations,
        # квартиры
        'apt_total': apt_total,
        'apt_sold': apt_sold,
        'apt_reserved': apt_reserved,
        'apt_barter': apt_barter,
        'apt_free': apt_free,
        # графики (JSON)
        'project_income_labels': json.dumps(project_income_labels, ensure_ascii=False),
        'project_income_data':   json.dumps(project_income_data),
        'block_income_labels':   json.dumps(block_income_labels, ensure_ascii=False),
        'block_income_data':     json.dumps(block_income_data),
        'expense_labels':        json.dumps(expense_labels, ensure_ascii=False),
        'expense_data':          json.dumps(expense_data),
        'monthly_labels':        json.dumps(monthly_labels, ensure_ascii=False),
        'monthly_income':        json.dumps(monthly_income),
        'monthly_expense':       json.dumps(monthly_expense),
    }
    return render(request, 'finances/dashboard.html', context)
import datetime
from django.utils.dateparse import parse_date
from django.http import HttpResponse

from openpyxl.utils import get_column_letter
from django.db.models import Q
from projects.models import Block
from .models import CommonCash, CashFlow



import openpyxl
from io import BytesIO
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from django.http import HttpResponse
from datetime import datetime

def export_cash_flows_to_excel(cash_flows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Flows"

    # 🔹 Заголовки
    headers = ["Дата", "Тип", "Сумма", "Описание", "Блок", "Создал"]
    ws.append(headers)

    # 🔹 Стили заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 🔹 Данные
    for flow in cash_flows:
        ws.append([
            flow.date.strftime("%d.%m.%Y %H:%M"),
            "Приход" if flow.flow_type == "income" else "Расход",
            float(flow.amount),
            flow.description,
            flow.block.name if flow.block else "—",            # Название блока
            flow.created_by.get_full_name() or flow.created_by.username,  # ФИО или username
        ])

    # 🔹 Автоширина колонок
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # 🔹 Итоги
    total_income = sum(f.amount for f in cash_flows if f.flow_type == "income")
    total_expense = sum(f.amount for f in cash_flows if f.flow_type == "expense")
    net_balance = total_income - total_expense

    ws.append([])
    ws.append(["", "ИТОГО:"])
    ws.append(["", "Общий приход", float(total_income)])
    ws.append(["", "Общий расход", float(total_expense)])
    ws.append(["", "Разница", float(net_balance)])

    # 🔹 Запись в память
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # 🔹 Имя файла с датой
    filename = f"cash_flows_{datetime.now().strftime('%d-%m-%Y')}.xlsx"

    # 🔹 HTTP-ответ
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response




@login_required
def allocations_list(request):
    allocations = Allocation.objects.all().order_by('-date')
    
    context = {
        'allocations': allocations,
    }
    return render(request, 'finances/allocations_list.html', context)

@login_required
def expenses_list(request):
    expenses = Expense.objects.all().order_by('-date')
    
    context = {
        'expenses': expenses,
    }
    return render(request, 'finances/expenses_list.html', context)

@login_required
def sales_list(request):
    sales = Sale.objects.all().order_by('-date')
    
    context = {
        'sales': sales,
    }
    return render(request, 'finances/sales_list.html', context)


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from .forms import AllocationForm
from .models import CommonCash, CashFlow, Allocation
from projects.models import EstimateItem


# views.py
from projects.models import Block
# finances/views.py (часть)
from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import AllocationForm
from .models import CommonCash, CashFlow, Allocation
from projects.models import Block, EstimateItem
from django.shortcuts import render, get_object_or_404, redirect
from .models import Block, EstimateItem, CommonCash, Allocation
from .forms import AllocationForm



def get_estimate_items(request):
    block_id = request.GET.get("block_id")
    items = EstimateItem.objects.filter(block_id=block_id).values("id", "name") if block_id else []
    return JsonResponse(list(items), safe=False)

def get_estimate_items3(request):
    """
    Возвращает JSON список позиций сметы.
    Параметры GET:
      - q: строка поиска (опционально)
      - block_id: id блока (опционально)
    """
    q = request.GET.get('q', '').strip()
    block_id = request.GET.get('block_id')

    qs = EstimateItem.objects.select_related('block', 'category').all()

    if block_id:
        qs = qs.filter(block_id=block_id)

    if q:
        qs = qs.filter(
            Q(name__icontains=q) |
            Q(block__name__icontains=q) |
            Q(category__name__icontains=q)
        )

    # лимитируем результат
    items = []
    for it in qs[:200]:
        display = f"{it.block.name} — {it.category.name} — {it.name}"
        items.append({'id': it.id, 'text': display})

    return JsonResponse({'results': items})




from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .forms import LoanForm, LoanPaymentForm
from .models import Loan, LoanPayment

@login_required
def loans_list(request):
    loans = Loan.objects.all().order_by('-issued_date')
    
    # Фильтрация по типу займа
    loan_type = request.GET.get('type')
    if loan_type in ['given', 'taken']:
        loans = loans.filter(loan_type=loan_type)
    
    # Фильтрация по статусу
    status = request.GET.get('status')
    if status in ['active', 'repaid', 'overdue']:
        loans = loans.filter(status=status)
    
    context = {
        'loans': loans,
        'current_type': loan_type,
        'current_status': status,
    }
    return render(request, 'finances/loans_list.html', context)

@login_required
def loan_detail(request, loan_id):
    loan = get_object_or_404(Loan, id=loan_id)
    payments = loan.payments.all().order_by('-payment_date')
    
    context = {
        'loan': loan,
        'payments': payments,
    }
    return render(request, 'finances/loan_detail.html', context)

@login_required
def create_loan(request):
    common_cash = CommonCash.objects.first()
    
    if request.method == 'POST':
        form = LoanForm(request.POST)
        if form.is_valid():
            loan = form.save(commit=False)
            loan.common_cash = common_cash
            loan.created_by = request.user
            
            # Проверяем достаточно ли средств для выдачи займа
            if loan.loan_type == 'given' and common_cash.balance < loan.amount:
                messages.error(request, 'Недостаточно средств в Общаге для выдачи займа')
                return render(request, 'finances/loan_form.html', {'form': form})
            
            loan.save()
            messages.success(request, 'Займ успешно создан')
            return redirect('finances:loans_list')
    else:
        form = LoanForm()
    
    context = {
        'form': form,
        'common_cash': common_cash,
    }
    return render(request, 'finances/loan_form.html', context)

@login_required
def add_loan_payment(request, loan_id):
    loan = get_object_or_404(Loan, id=loan_id)
    common_cash = loan.common_cash
    
    if request.method == 'POST':
        form = LoanPaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.loan = loan
            payment.created_by = request.user
            
            # Проверяем достаточно ли средств для возврата полученного займа
            if loan.loan_type == 'taken' and common_cash.balance < payment.amount:
                messages.error(request, 'Недостаточно средств в Общаге для возврата займа')
                return render(request, 'finances/loan_payment_form.html', {'form': form, 'loan': loan})
            
            # Проверяем не превышает ли платеж остаток по займу
            if payment.amount > loan.remaining_amount:
                messages.error(request, f'Сумма платежа не может превышать остаток по займу: {loan.remaining_amount}')
                return render(request, 'finances/loan_payment_form.html', {'form': form, 'loan': loan})
            
            payment.save()
            messages.success(request, 'Платеж по займу успешно добавлен')
            return redirect('finances:loan_detail', loan_id=loan.id)
    else:
        form = LoanPaymentForm()
    
    context = {
        'form': form,
        'loan': loan,
    }
    return render(request, 'finances/loan_payment_form.html', context)










@login_required
def car_list(request):
    cars = WarehouseCar.objects.all()
    return render(request, "finances/car_list.html", {"cars": cars})


@login_required
def car_purchase(request):
    if request.method == "POST":
        form = CarPurchaseForm(request.POST)
        if form.is_valid():
            car = form.save(commit=False)
            car.common_cash = CommonCash.objects.first()
            car.created_by = request.user
            car.save()
            return redirect("finances:car_list")
    else:
        form = CarPurchaseForm()
    return render(request, "finances/car_purchase.html", {"form": form})


@login_required
def car_sale(request, pk):
    car = get_object_or_404(WarehouseCar, pk=pk, status="available")
    if request.method == "POST":
        form = CarSaleForm(request.POST, instance=car)
        if form.is_valid():
            car = form.save(commit=False)
            car.status = "sold"
            car.sale_date = timezone.now()
            car._mark_as_sold = True  # триггер для сохранения даты
            car.save()
            return redirect("finances:car_list")
    else:
        form = CarSaleForm(instance=car)
    return render(request, "finances/car_sale.html", {"form": form, "car": car})


from django.contrib.auth.decorators import user_passes_test

def superuser_required(view_func):
    """Декоратор для проверки, что пользователь - суперпользователь"""
    decorated_view_func = user_passes_test(
        lambda u: u.is_active and u.is_superuser,
        login_url='/admin/login/'
    )(view_func)
    return decorated_view_func

@login_required
@superuser_required
def delete_allocation(request, allocation_id):
    allocation = get_object_or_404(Allocation, id=allocation_id)
    
    if request.method == 'POST':
        # Сохраняем информацию для сообщения
        allocation_info = f"{allocation.amount} сом на {allocation.estimate_item}"
        
        # Удаляем выделение (в методе delete уже реализована логика возврата средств)
        allocation.delete()
        
        messages.success(request, f'Выделение средств {allocation_info} успешно удалено. Средства возвращены в Общаг.')
        return redirect('finances:common_cash_detail2')
    
    context = {
        'allocation': allocation,
    }
    return render(request, 'finances/delete_allocation_confirm.html', context)



from django.contrib import messages
from django.contrib.auth.decorators import login_required

# если уже импортировано — второй раз не нужно
# from .models import CommonCash, Block, EstimateItem
# from .forms import AllocationForm


@login_required
def allocation_create(request):
    common_cash = CommonCash.objects.first()
    blocks = Block.objects.all()

    block_id = request.GET.get("block")
    search_query = request.GET.get("q", "").strip()  # текст поиска по позициям сметы

    form = None
    estimate_items = EstimateItem.objects.none()  # по умолчанию пусто

    if block_id:
        # базовый фильтр по выбранному блоку
        estimate_items = EstimateItem.objects.filter(block_id=block_id)

        # если введён текст поиска — дополнительно фильтруем
        if search_query:
            # ⚠️ предполагаю, что в EstimateItem есть поле name
            # при необходимости поменяй на title / description и т.п.
            estimate_items = estimate_items.filter(name__icontains=search_query)

        if request.method == "POST":
            form = AllocationForm(request.POST)
            # ограничиваем выбор позиций только отфильтрованным queryset'ом
            form.fields["estimate_item"].queryset = estimate_items

            if form.is_valid():
                try:
                    allocation = form.save(commit=False)

                    # проверяем достаточно ли средств
                    if common_cash.balance < allocation.amount:
                        messages.error(request, "Недостаточно средств в Общаге")
                        return render(
                            request,
                            "finances/allocation_create.html",
                            {
                                "form": form,
                                "common_cash": common_cash,
                                "blocks": blocks,
                                "block_id": int(block_id),
                                "search_query": search_query,
                            },
                        )

                    allocation.common_cash = common_cash
                    allocation.created_by = request.user
                    allocation.save()

                    messages.success(
                        request,
                        f"Средства в размере {allocation.amount} сом успешно выделены!",
                    )
                    return redirect("finances:common_cash_detail")

                except Exception as e:
                    messages.error(
                        request,
                        f"Ошибка при выделении средств: {str(e)}",
                    )
        else:
            form = AllocationForm()
            form.fields["estimate_item"].queryset = estimate_items

    context = {
        "common_cash": common_cash,
        "blocks": blocks,
        "block_id": int(block_id) if block_id else None,
        "form": form,
        "search_query": search_query,
    }
    return render(request, "finances/allocation_create.html", context)

# @login_required
# def allocation_create(request):
#     common_cash = CommonCash.objects.first()
#     blocks = Block.objects.all()
#     block_id = request.GET.get("block")

#     form = None
#     if block_id:
#         estimate_items = EstimateItem.objects.filter(block_id=block_id)
#         if request.method == "POST":
#             form = AllocationForm(request.POST)
#             form.fields["estimate_item"].queryset = estimate_items
#             if form.is_valid():
#                 try:
#                     allocation = form.save(commit=False)
                    
#                     # Проверяем достаточно ли средств в Общаге
#                     if common_cash.balance < allocation.amount:
#                         messages.error(request, 'Недостаточно средств в Общаге')
#                         return render(request, 'finances/allocation_create.html', {
#                             'form': form, 
#                             'common_cash': common_cash, 
#                             'blocks': blocks, 
#                             'block_id': block_id
#                         })

#                     # Создаем запись в движении денег
#                     # cash_flow = CashFlow.objects.create(
#                     #     common_cash=common_cash,
#                     #     flow_type='expense',
#                     #     amount=allocation.amount,
#                     #     block=allocation.estimate_item.block,
#                     #     description=f"Выделение средств: {allocation.description}",
#                     #     created_by=request.user
#                     # )

#                     # Сохраняем выделение средств
#                     allocation.common_cash = common_cash
#                     allocation.created_by = request.user
#                     allocation.save()

#                     messages.success(request, f'Средства в размере {allocation.amount} сом успешно выделены!')
#                     return redirect("finances:common_cash_detail")
                    
#                 except Exception as e:
#                     messages.error(request, f'Ошибка при выделении средств: {str(e)}')
#         else:
#             form = AllocationForm()
#             form.fields["estimate_item"].queryset = estimate_items

#     context = {
#         "common_cash": common_cash,
#         "blocks": blocks,
#         "block_id": int(block_id) if block_id else None,
#         "form": form,
#     }
#     return render(request, "finances/allocation_create.html", context)